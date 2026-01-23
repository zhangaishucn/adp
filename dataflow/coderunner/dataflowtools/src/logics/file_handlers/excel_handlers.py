#!/usr/bin/python3
# -*- coding:utf-8 -*-

import xlrd
import openpyxl
from xlutils.copy import copy
from typing import List, Any
from .base_handler import BaseFileHandler
from models.file_operator import FileOperatorRequest
from common.configs import FileOperatorType, InsertType, DataType


class XlsxHandler(BaseFileHandler):
    """XLSX文件处理器"""

    SUPPORT_NEW_TYPES = [DataType.NEW_ROW.value, DataType.NEW_COL.value]
    SUPPORT_INSERT_TYPES = [
        InsertType.APPEND.value,
        InsertType.APPEND_BEFORE.value,
        InsertType.APPEND_AFTER.value,
        InsertType.COVER.value,
    ]

    def get_file_extension(self) -> str:
        return '.xlsx'

    def needs_temp_file(self) -> bool:
        return True

    def supports_operation(self, request: FileOperatorRequest) -> bool:
        """检查是否支持该操作"""
        # XLSX支持行列操作，不支持内容插入
        if request.new_type in self.SUPPORT_NEW_TYPES and request.insert_type in self.SUPPORT_INSERT_TYPES:
            return True
        return False

    def _do_update(self, source_path: str, target_path: str, request: FileOperatorRequest) -> None:
        """
        统一的更新接口
        
        Args:
            data_source: BytesIO对象
            request: 更新请求
            
        Returns:
            更新后的数据
        """

        try:
            wb = openpyxl.load_workbook(source_path)
        except FileNotFoundError:
            if request.op == FileOperatorType.CREATE.value:
                wb = openpyxl.Workbook()
            else:
                raise
        sheet = wb.worksheets[0]

        # 根据操作类型分发
        if request.new_type == DataType.NEW_ROW.value:
            self._process_row_insertion(
                sheet, 
                request.insert_type, 
                request.insert_pos or 1,
                request.content
            )
        elif request.new_type == DataType.NEW_COL.value:
            self._process_col_insertion(
                sheet,
                request.insert_type,
                request.insert_pos or 1,
                request.content
            )

        wb.save(target_path)
        wb.close()

    def _process_row_insertion(self, sheet, insert_type: str, insert_pos: int, insert_data: List[Any]):
        """
        处理行插入逻辑
        
        Args:
            sheet: openpyxl.worksheet.Worksheet对象
            insert_type: 插入类型append|append_before|append_after|cover
            insert_pos: 插入位置
            insert_data: 需要插入的数据
        """
        if insert_type == InsertType.APPEND.value:
            sheet.append(insert_data)
        elif insert_type == InsertType.APPEND_BEFORE.value:
            sheet.insert_rows(insert_pos)
            self._write_row_data(sheet, insert_pos, insert_data)
        elif insert_type == InsertType.APPEND_AFTER.value:
            insert_pos += 1
            sheet.insert_rows(insert_pos)
            self._write_row_data(sheet, insert_pos, insert_data)
        elif insert_type == InsertType.COVER.value:
            self._write_row_data(sheet, insert_pos, insert_data)

    def _process_col_insertion(self, sheet, insert_type: str, insert_pos: int, insert_data: List[Any]):
        """
        处理列插入逻辑
        
        Args:
            sheet: openpyxl.worksheet.Worksheet对象
            insert_type: 插入类型append|append_before|append_after|cover
            insert_pos: 插入位置
            insert_data: 需要插入的数据
        """
        if insert_type == InsertType.APPEND.value:
            last_column_index = sheet.max_column + 1
            if sheet.max_column == 1 and sheet.max_row == 1 and sheet.cell(1, 1).value is None:
                last_column_index = sheet.max_column
            self._write_col_data(sheet, last_column_index, insert_data)
        elif insert_type == InsertType.APPEND_BEFORE.value:
            sheet.insert_cols(insert_pos)
            self._write_col_data(sheet, insert_pos, insert_data)
        elif insert_type == InsertType.APPEND_AFTER.value:
            insert_pos += 1
            sheet.insert_cols(insert_pos)
            self._write_col_data(sheet, insert_pos, insert_data)
        elif insert_type == InsertType.COVER.value:
            self._write_col_data(sheet, insert_pos, insert_data)

    @staticmethod
    def _write_row_data(sheet, row_pos: int, data: List[Any]):
        """写入行数据"""
        for col_num, value in enumerate(data, start=1):
            sheet.cell(row=row_pos, column=col_num, value=value)

    @staticmethod
    def _write_col_data(sheet, col_pos: int, data: List[Any]):
        """写入列数据"""
        for row_num, value in enumerate(data, start=1):
            sheet.cell(row=row_num, column=col_pos, value=value)


class XlsHandler(BaseFileHandler):
    """XLS文件处理器"""

    SUPPORT_NEW_TYPES = [DataType.NEW_ROW.value, DataType.NEW_COL.value]
    SUPPORT_INSERT_TYPES = [
        InsertType.APPEND.value,
        InsertType.APPEND_BEFORE.value,
        InsertType.APPEND_AFTER.value,
        InsertType.COVER.value,
    ]
    
    def get_file_extension(self) -> str:
        return '.xls'
    
    def needs_temp_file(self) -> bool:
        return True
    
    def supports_operation(self, request: FileOperatorRequest) -> bool:
        """检查是否支持该操作"""
        # XLS支持行列操作，不支持内容插入
        if request.new_type in self.SUPPORT_NEW_TYPES and request.insert_type in self.SUPPORT_INSERT_TYPES:
            return True
        return False
    
    def _do_update(self, source_path: str, target_path: str, request: FileOperatorRequest) -> None:
        """
        统一的更新接口
        
        Args:
            file_path: 临时文件路径
            request: 更新请求
            
        Returns:
            更新后的数据
        """
        
        xlrd_workbook = xlrd.open_workbook(source_path, formatting_info=True)
        xlrd_sheet = xlrd_workbook.sheet_by_index(0)

        # 创建副本
        xlwt_workbook = copy(xlrd_workbook)
        xlwt_sheet = xlwt_workbook.get_sheet(0)
        
        # 根据操作类型分发
        if request.new_type == DataType.NEW_ROW.value:
            self._process_row_insertion(
                xlrd_sheet,
                xlwt_sheet,
                request.insert_type,
                request.insert_pos or 1,
                request.content
            )
        elif request.new_type == DataType.NEW_COL.value:
            self._process_col_insertion(
                xlrd_sheet,
                xlwt_sheet,
                request.insert_type,
                request.insert_pos or 1,
                request.content
            )
        
        xlwt_workbook.save(target_path)
        xlrd_workbook.release_resources()
    
    def _process_row_insertion(self, xlrd_sheet, xlwt_sheet, insert_type: str, insert_pos: int, insert_data: List[Any]):
        """处理行插入逻辑"""
        if insert_type == InsertType.APPEND.value:
            for col_index, value in enumerate(insert_data):
                xlwt_sheet.write(xlrd_sheet.nrows, col_index, value)
        elif insert_type in [InsertType.APPEND_BEFORE.value, InsertType.APPEND_AFTER.value]:
            actual_pos = insert_pos - 1 if insert_type == InsertType.APPEND_BEFORE.value else insert_pos
            self._insert_row_at_position(xlrd_sheet, xlwt_sheet, actual_pos, insert_data)
        elif insert_type == InsertType.COVER.value:
            for col_index, value in enumerate(insert_data):
                xlwt_sheet.write(insert_pos - 1, col_index, value)
    
    def _process_col_insertion(self, xlrd_sheet, xlwt_sheet, insert_type: str, insert_pos: int, insert_data: List[Any]):
        """处理列插入逻辑"""
        if insert_type == InsertType.APPEND.value:
            for row_index, value in enumerate(insert_data):
                xlwt_sheet.write(row_index, xlrd_sheet.ncols, value)
        elif insert_type in [InsertType.APPEND_BEFORE.value, InsertType.APPEND_AFTER.value]:
            actual_pos = insert_pos - 1 if insert_type == InsertType.APPEND_BEFORE.value else insert_pos
            self._insert_col_at_position(xlrd_sheet, xlwt_sheet, actual_pos, insert_data)
        elif insert_type == InsertType.COVER.value:
            for row_index, value in enumerate(insert_data):
                xlwt_sheet.write(row_index, insert_pos - 1, value)
    
    def _insert_row_at_position(self, xlrd_sheet, xlwt_sheet, insert_pos: int, insert_data: List[Any]):
        """在指定位置插入行"""
        for row_index in range(xlrd_sheet.nrows):
            row_data = xlrd_sheet.row_values(row_index)
            
            if row_index == insert_pos:
                for col_index, value in enumerate(insert_data):
                    xlwt_sheet.write(row_index, col_index, value)
            
            actual_row = row_index + 1 if row_index >= insert_pos else row_index
            for col_index, value in enumerate(row_data):
                xlwt_sheet.write(actual_row, col_index, value)
        
        if xlrd_sheet.nrows <= insert_pos:
            self._fill_blank_rows(xlwt_sheet, xlrd_sheet.nrows, insert_pos, insert_data)
    
    def _insert_col_at_position(self, xlrd_sheet, xlwt_sheet, insert_pos: int, insert_data: List[Any]):
        """在指定位置插入列"""
        for col_index in range(xlrd_sheet.ncols):
            col_data = xlrd_sheet.col_values(col_index)
            
            if col_index == insert_pos:
                for row_index, value in enumerate(insert_data):
                    xlwt_sheet.write(row_index, col_index, value)
            
            actual_col = col_index + 1 if col_index >= insert_pos else col_index
            for row_index, value in enumerate(col_data):
                xlwt_sheet.write(row_index, actual_col, value)
        
        if xlrd_sheet.ncols <= insert_pos:
            self._fill_blank_cols(xlwt_sheet, xlrd_sheet.ncols, insert_pos, insert_data)
    
    @staticmethod
    def _fill_blank_rows(xlwt_sheet, last_row_index: int, insert_pos: int, insert_data: List[Any]):
        """填充空白行"""
        add_blank_num = insert_pos - last_row_index
        for i in range(add_blank_num):
            xlwt_sheet.write(last_row_index + i, 0, "")
        for col_index, value in enumerate(insert_data):
            xlwt_sheet.write(last_row_index + add_blank_num, col_index, value)
    
    @staticmethod
    def _fill_blank_cols(xlwt_sheet, last_col_index: int, insert_pos: int, insert_data: List[Any]):
        """填充空白列"""
        add_blank_num = insert_pos - last_col_index
        for i in range(add_blank_num):
            xlwt_sheet.write(0, last_col_index + i, "")
        for row_index, value in enumerate(insert_data):
            xlwt_sheet.write(row_index, last_col_index + add_blank_num, value)
